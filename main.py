
from random import randint
from dotenv import load_dotenv
from time import sleep
from multiprocessing import Pool
from concurrent.futures import ThreadPoolExecutor
from bs4 import BeautifulSoup
import os
import asyncio
import concurrent.futures
import requests
import time
import openpyxl
from aiohttp import ClientSession
import re


 
excel = openpyxl.load_workbook("data.xlsx", data_only=True)
session = requests.Session()


def getData(index): 
    rand_value = randint(10, 30)
    print(rand_value*0.1)
    sleep(rand_value*0.1)
    url = 'https://pcmap.place.naver.com/restaurant/' + str(index)
    html = session.get(url,headers={'Accept':'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9','User-Agent':'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/103.0.0.0 Safari/537.36'})
    print(html.status_code)
    while(html.status_code!=200):
        rand_value = randint(10, 30)
        print(rand_value*0.1)
        print(html.status_code)
        sleep(rand_value*0.1)
        print('다시하는 중')
        html = session.get(url,headers={'Accept':'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9','User-Agent':'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/103.0.0.0 Safari/537.36'})

    if html.text.find("og:title")!=-1:
        soup = BeautifulSoup(html.content,'html.parser',from_encoding='utf-8')
        if len(soup(text=re.compile('_3ocDE')))>0:
            return soup.select('span._3ocDE')[0].text
        else:
            return None
    else:
        return None


def incrementID():
    
    with open(".env",'r') as f: #open a file in the same folder
        currentId = f.readlines() 
    b = int(currentId[0].split('=')[1])                    #get integer at first position
    b = b+1                          #increment
    with open(".env",'w') as f: #open same file
        f.write('CURRENT_ID='+str(b))
    
def loadCurrentIdFromEnv():
    return os.getenv('CURRENT_ID')
    
def addDataToExcel(data):
    # 엑셀파일 쓰기
    sheet = excel.worksheets[0]
    titles = sheet['A']
    isAdded = False
    for index in range(0, len(titles)):
        if titles[index].value == data:
            sheet.cell(row=index+1,column=2).value = str(int(sheet.cell(row=index+1,column=2).value) + 1)
            isAdded = True
            break
        
    if isAdded != True :
        sheet.append([data,'1'])
    
    # excel.worksheets[0] = sheet
    excel.save(filename="data.xlsx")


def main(value,):
    print(value)
    addData = getData(value)
    print(addData)
    if(addData!=None):
        addDataToExcel(addData,)
        
        
def main_exec(value,):
    thread_list=list()
    with ThreadPoolExecutor(max_workers=9) as executor:
        for index in range(value*20000,(value+1)*20000):
            thread_list.append(executor.submit(main(index,)))
        # for execution in concurrent.futures.as_completed(thread_list):
        #     incrementID()
            




if __name__ == "__main__":
    load_dotenv()
    currentId = loadCurrentIdFromEnv()
    addData = ''
    start_time = time.time()
    with Pool(processes=10) as pool:  
        pool.map(main_exec,[0,1,2,3,4],)
    


        
    print("--- elapsed time %s seconds ---" % (time.time() - start_time))